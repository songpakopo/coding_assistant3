from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os
import math

app = Flask(__name__)

# Excel file path
EXCEL_FILE = 'stock.xlsx'
PER_PAGE = 5

def init_excel():
    """Create the Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Headers matching the stock data
        sheet.append(['ID', '종목코드', '회사명', '현재가', '거래량', '예측'])
        # Add some initial data for demonstration
        sheet.append([2, '005930', '삼성전자', 70000, 10000000, 1])
        sheet.append([3, '000660', 'SK하이닉스', 130000, 5000000, 0])
        workbook.save(EXCEL_FILE)

@app.route('/')
def index():
    """Display records from the Excel file with pagination, sorting, and searching."""
    init_excel()
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    
    data = []
    # Read data starting from the second row
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if any(cell is not None for cell in row): # Check if row is not entirely empty
            data.append({
                'id': row_idx,
                'code': row[1],
                'name': row[2],
                'price': row[3],
                'volume': row[4],
                'prediction': row[5]
            })

    # Search
    search_query = request.args.get('search', '')
    if search_query:
        data = [item for item in data if search_query.lower() in str(item['name']).lower() or search_query.lower() in str(item['code']).lower()]

    # Sorting
    sort_by = request.args.get('sort_by', 'name')
    sort_order = request.args.get('sort_order', 'asc')
    
    if sort_by and sort_by in data[0]:
        data.sort(key=lambda x: x[sort_by], reverse=sort_order == 'desc')

    # Pagination
    page = request.args.get('page', 1, type=int)
    total_items = len(data)
    total_pages = math.ceil(total_items / PER_PAGE)
    start = (page - 1) * PER_PAGE
    end = start + PER_PAGE
    paginated_data = data[start:end]

    return render_template('index.html', 
                           data=paginated_data,
                           page=page,
                           total_pages=total_pages,
                           sort_by=sort_by,
                           sort_order=sort_order,
                           search_query=search_query)

@app.route('/add', methods=['GET', 'POST'])
def add():
    """Add a new record to the Excel file."""
    if request.method == 'POST':
        code = request.form['code']
        name = request.form['name']
        price = request.form['price']
        volume = request.form['volume']
        prediction = request.form['prediction']

        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        
        # Find the next available row ID
        next_id = sheet.max_row + 1
        
        sheet.append([next_id, code, name, price, volume, prediction])
        workbook.save(EXCEL_FILE)
        return redirect(url_for('index'))
    return render_template('add.html')

@app.route('/edit/<int:row_id>', methods=['GET', 'POST'])
def edit(row_id):
    """Edit a record in the Excel file."""
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    if request.method == 'POST':
        sheet.cell(row=row_id, column=2, value=request.form['code'])
        sheet.cell(row=row_id, column=3, value=request.form['name'])
        sheet.cell(row=row_id, column=4, value=request.form['price'])
        sheet.cell(row=row_id, column=5, value=request.form['volume'])
        sheet.cell(row=row_id, column=6, value=request.form['prediction'])
        workbook.save(EXCEL_FILE)
        return redirect(url_for('index'))

    # Get current data for the form
    row_values = sheet[row_id]
    data = {
        'id': row_id,
        'code': row_values[1].value,
        'name': row_values[2].value,
        'price': row_values[3].value,
        'volume': row_values[4].value,
        'prediction': row_values[5].value
    }
    return render_template('edit.html', data=data)

@app.route('/delete/<int:row_id>')
def delete(row_id):
    """Delete a record from the Excel file by clearing its content."""
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    
    # Clear the row content to avoid shifting row IDs
    for col_idx in range(1, sheet.max_column + 1):
        sheet.cell(row=row_id, column=col_idx, value=None)
        
    workbook.save(EXCEL_FILE)
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)